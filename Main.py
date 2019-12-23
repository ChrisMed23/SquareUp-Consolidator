from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from bs4 import BeautifulSoup as soup
from urllib.request import urlopen as uReq




my_url = input("Enter Squareup url: ")
print("Parsing data...")
Prices = []
Products = []


uClient = uReq(my_url)
page_html = uClient.read()
uClient.close()
page_soup = soup(page_html, "html.parser")

#get filename
title = page_soup.find("td", {"merchant-header__name"})
title = title.text
#worksheet name has max of 31 chars
title = title[:30]

#get items in table
products = page_soup.findAll("div", {"class", "p item-name"})
for product in products:
    Products.append(product.text)


#get prices for items
product_prices = page_soup.findAll("td", {"class": "half-col-right"})
for price in product_prices:
    Prices.append((price.text).strip("$"))

#write to book
try:
    # load file if already created and create a new worksheet
    wb = load_workbook("Purchases.xlsx")
    ws = wb.create_sheet(title)
    print("Writing to workbook.")
except Exception as e:
    # creates a workbook is first time running
    wb = Workbook()
    ws = wb.active
    ws.title = title
    print("Creating Workbook.")


ws['A1'] = "Products"
ws['A1'].font = Font(bold=True, size=14)
ws['B1'] = "Prices"
ws['B1'].font = Font(bold=True, size=14)

for i in range(0, len(Products)):
    ws['A' + str(i+2)] = Products[i]


ws['A' + str(len(Products)+2)] = "Subtotal"
ws['A' + str(len(Products)+2)].font = Font(bold=True, size=14)
ws['A' + str(len(Products)+3)] = "Tax"
ws['A' + str(len(Products)+3)].font = Font(bold=True, size=14)
ws['A' + str(len(Products)+4)] = "Total"
ws['A' + str(len(Products)+4)].font = Font(bold=True, size=14)

for i in range(0, len(Prices)-2):
    ws['B' + str(i+2)] = float(Prices[i])

wb.save("Purchases.xlsx")
wb.close()

print("Done!")
